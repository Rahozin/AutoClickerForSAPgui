"""Auto R3 docs"""

from time import sleep
from datetime import datetime, timedelta
import sys
import pyautogui
import keyboard
import pyperclip
import arrow
import openpyxl
import win32api


def lang_en():
    """Layout to en"""
    while pyautogui.locateOnScreen(PICS_PATH + 'lang_en.png',
                                   region=(1800, 1030, 120, 50), confidence=0.8) is None:
        pyautogui.hotkey("ctrl", "shift")
        sleep(SLEEPTIME)


def lang_en_manualy():
    """Layout to en manualy"""
    pyautogui.keyDown('win')
    key('space')
    sleep(SLEEPTIME)
    pyautogui.click(x=1770, y=540)
    pyautogui.keyUp('win')
    sleep(SLEEPTIME)


def check_layout():
    """Exit if layout isn't ENG"""
    layout = win32api.GetKeyboardLayout()
    if layout != 67699721:
        MESSAGE3 = 'Зміни розкладку клавіатури на "ENG"\n і перезапусти програму.'
        missing3 = pyautogui.alert(
            text=MESSAGE3, title='Неправильна розкладка клавіатури!', button='Ok')
        if missing3 == 'Ok':
            sys.exit()


def quick_click(clicks, pic_name):
    """Clicking exactly when pic is appear + interapt"""
    if keyboard.is_pressed('ctrl') is False:
        time_start = datetime.now()
        time_finish = time_start + timedelta(seconds=10)
        go_click = True
        while go_click:
            target = pyautogui.locateOnScreen(
                PICS_PATH + pic_name, confidence=0.8)
            if target:
                if clicks == 1:
                    pyautogui.click(target)
                    print('1 click')
                if clicks == 2:
                    pyautogui.doubleClick(target)
                    print('2 click')
                break
            if time_start >= time_finish:
                message = 'Не знаходжу кнопку: ' + pic_name + \
                    '.\n' + 'Можна зробити крок вручну і продовжити'
                missing = pyautogui.confirm(
                    text=message, title='Щось пішло не так!', buttons=['Продовжити', 'Зупинити'])
                if missing == 'Зупинити':
                    sys.exit()
                else:
                    go_click = False
            time_start = datetime.now()
            sleep(0.2)
    else:
        pyautogui.alert('Програма була зупинена!!')
        sys.exit()


def check_done(pic_name, i):
    """Checking if event was done"""

    time_start = datetime.now()
    time_finish = time_start + timedelta(seconds=5)
    check = True
    while check:
        target = pyautogui.locateOnScreen(
            PICS_PATH_CHECK + pic_name, confidence=0.95)
        if target:
            break
        if time_start >= time_finish:
            print('Була пропущена лінія.' + str(i+1) + 'Повторне внесення')
            i = i - 1
            check = False
        time_start = datetime.now()
        sleep(SLEEPTIME)
    return i


def click(pic_name):
    """Simplyfying common functions"""
    clicks = 1
    quick_click(clicks, pic_name)


def click2(pic_name):
    """Simplyfying common functions"""
    clicks = 2
    quick_click(clicks, pic_name)


def write(string):
    """Simplyfying common functions"""
    keyboard.write(string)
    print('text')


def key(string):
    """Simplyfying common functions"""
    pyautogui.press(string)


def hotkey(string1, string2):
    """Simplyfying common functions + interapt"""
    if keyboard.is_pressed('ctrl') is False:
        pyautogui.hotkey(string1, string2)
    else:
        pyautogui.alert('Програма була зупинена!')
        sys.exit()


def press_and_release(string):
    """Simplyfying common functions + interapt"""
    if keyboard.is_pressed('ctrl') is False:
        keyboard.press_and_release(string)
    else:
        pyautogui.alert('Програма була зупинена!')
        sys.exit()


def login():
    """loging in"""
    click2('r3.png')
    click('login.png')
    write('vtv3')
    click('password.png')
    write('1234')
    key('enter')
    while pyautogui.locateOnScreen(PICS_PATH + 'new_order.png') is None:
        if pyautogui.locateOnScreen(PICS_PATH + 'multiple_registration.png'):
            click('multiple_registration.png')
            key('enter')


def create_order():
    """creating an order"""
    click2('new_order.png')
    click('sector.png')
    key('backspace')
    write(SECTOR)
    key('enter')


def fill_order():
    """filling an order"""
    click('EDRPOU.png')
    write(edrpou)
    key('enter')
    click('choose_order.png')
    key('enter')
    click('material.png')
    sleep(SLEEPTIME*2)

    output1 = ''
    output2 = ''
    output3 = ''
    output4 = ''
    output5 = ''
    output6 = ''
    output7 = ''

    for i in range(1, positions_quantity+1):
        material = MATERIAL09
        if i in lines_list_k1:
            material = MATERIAL1

        if i in range(1, 18):
            TOTAL_OUTPUT_QUANTITY = 1
            output1 = output1 + material + '\t' + \
                str(1) + '\t'*4 + str(50150) + '\t' + '\n'

        if i in range(18, 34):
            TOTAL_OUTPUT_QUANTITY = 2
            output2 = output2 + material + '\t' + \
                str(1) + '\t'*4 + str(50150) + '\t' + '\n'

        if i in range(34, 50):
            TOTAL_OUTPUT_QUANTITY = 3
            output3 = output3 + material + '\t' + \
                str(1) + '\t'*4 + str(50150) + '\t' + '\n'

        if i in range(50, 66):
            TOTAL_OUTPUT_QUANTITY = 4
            output4 = output4 + material + '\t' + \
                str(1) + '\t'*4 + str(50150) + '\t' + '\n'

        if i in range(66, 82):
            TOTAL_OUTPUT_QUANTITY = 5
            output5 = output5 + material + '\t' + \
                str(1) + '\t'*4 + str(50150) + '\t' + '\n'

        if i in range(82, 98):
            TOTAL_OUTPUT_QUANTITY = 6
            output6 = output6 + material + '\t' + \
                str(1) + '\t'*4 + str(50150) + '\t' + '\n'

        if i in range(98, 113):
            TOTAL_OUTPUT_QUANTITY = 7
            output7 = output7 + material + '\t' + \
                str(1) + '\t'*4 + str(50150) + '\t' + '\n'

    for n in range(1, TOTAL_OUTPUT_QUANTITY+1):
        output = locals().get("output" + str(n))
        pyperclip.copy(output)
        press_and_release('ctrl + v')
        pyautogui.press("enter", presses=19, interval=1)
        sleep(SLEEPTIME*2)


def fill_work_type():
    """filling lines work type"""
    click2('work_type1.png')
    sleep(SLEEPTIME*2.5)
    hotkey("shift", "f5")
    for _ in range(positions_quantity):
        click('work_type2.png')
        write(WORK_TYPE)
        hotkey("shift", "f7")
        sleep(SLEEPTIME)


def fill_lines_names():
    """filling lines names"""
    pyautogui.click(pyautogui.locateOnScreen(PICS_PATH + 'line_name1.png',
                    confidence=0.8), clicks=10, interval=0.1)
    click('line_name2.png')
    sleep(SLEEPTIME*2.5)
    hotkey("shift", "f5")
    for i in range(positions_quantity):
        pyperclip.copy(lines_list[i])
        click('line_name3.png')
        press_and_release('ctrl + v')
        hotkey("shift", "f7")
        sleep(SLEEPTIME)


def fill_emails():
    """filling emails"""
    click('Email1.png')
    hotkey("shift", "f4")
    pyautogui.click(pyautogui.locateOnScreen(
        PICS_PATH + 'Email2.png', confidence=0.8), clicks=8, interval=0.1)
    click('Email3.png')
    click('Email4.png')
    pyautogui.dragRel(0, 350, 1)
    click('Email5.png')
    keyboard.write(EMAIL_POE)
    click('Email6.png')
    keyboard.write(cust_email)
    key("f3")


def calculate():
    """calculation"""
    click('select_all.png')
    click('calculation0.png')
    for _ in range(positions_quantity):
        click('calculation01.png')
        hotkey("ctrl", "f7")
        click('calculation1.png')
        click('calculation2.png')
        click('calculation3.png')
        click('calculation4.png')
        click('calculation5.png')


def factorize():
    """factoring"""
    factoring = pyautogui.confirm(text='Відфакторувати і зберегти?',
                                  title='', buttons=['Так', 'Ні'])
    if factoring == 'Ні':
        sys.exit()
    click('factoring1.png')
    click('factoring2.png')
    click('factoring3.png')
    click('factoring4.png')


def save_order_number():
    """saving order number"""
    click2('order_number1.png')
    sleep(SLEEPTIME)
    hotkey("ctrl", "a")
    hotkey("ctrl", "c")
    number = pyperclip.paste()
    with open(date + '_' + number + '_' + cust + '_' + str(positions_quantity) +
              ' л.txt', "w", encoding="UTF-8") as file:
        file.write(number)
        file.close()
    hotkey('shift', 'f3')
    return number


def make_contract(number):
    """making contract"""
    click2('make_contract1.png')
    sleep(SLEEPTIME*5)
    write(number)
    click('make_contract2.png')
    pyautogui.scroll(-5000)
    click('make_contract3.png')
    key('tab')
    hotkey('ctrl', 'a')
    write(SIGN1)
    key('tab')
    hotkey('ctrl', 'a')
    write(SIGN2)
    key('f8')
    sleep(SLEEPTIME*5)


def save_contract(number):
    """saving contract"""
    click('save_doc1.png')
    sleep(SLEEPTIME*5)
    click('save_doc2.png')
    sleep(SLEEPTIME*5)
    write('PDF!')
    key('enter')
    sleep(SLEEPTIME*10)
    pyautogui.hotkey('ctrl', 'shift', 's')
    sleep(SLEEPTIME*5)
    write(r'Y:\shared\ВРСП\Договора на ТУ\_Проекти договорів' + '\\' +
          date + '_' + number + '_' + cust + '_' + str(positions_quantity) + ' л')
    sleep(SLEEPTIME*10)
    click('save_doc3.png')


if __name__ == '__main__':

    SECTOR = "4"
    WORK_TYPE = "04"
    MATERIAL1 = "27001764"
    MATERIAL09 = "27001765"
    SLEEPTIME = 0.2
    EMAIL_POE = "vpr30@poe.pl.ua"
    PICS_PATH = 'pics/'
    PICS_PATH_CHECK = 'pics/check/'
    SIGN1 = 'Директор сервісного центру Гопка О.М.'
    SIGN2 = '.'

    date = arrow.now().format('YYMMDD')

    check_layout()

    # edrpou = '31801575'
    edrpou = pyautogui.prompt("ЕДРПОУ замовника: ")
    if edrpou is None:
        sys.exit()

    CUSTOMERS = 'customers.xlsx'
    wb = openpyxl.load_workbook(CUSTOMERS, read_only=True)
    ws = wb.active
    cust = ''
    cust_email = ''
    for row in ws.rows:
        for cell in row:
            if cell.value == int(edrpou):
                cust = ws.cell(cell.row, 3).value
                cust_email = ws.cell(cell.row, 4).value
    if cust_email == '' or cust == '':
        MESSAGE2 = 'Заповни дані про замовника в таблицю customers.xlsx'
        missing2 = pyautogui.alert(
            text=MESSAGE2, title='Немає даних про замовника!', button='Ok')
        if missing2 == 'Ok':
            sys.exit()
    wb.close()

    with open("lines_list.txt", "r", encoding="UTF-8") as data2:     # list of lines
        lines_data = data2.readlines()
        lines_list = list(
            filter(None, [element.replace('\n', '') for element in lines_data]))
        data2.close()
    positions_quantity = len(lines_list)

    lines_k1 = pyautogui.prompt(text='Напиши через кому по порядку, наприклад: 1,2,5,9',
                                title='Номера ліній з Кз = 1',
                                default='')  # list of lines with k=1
    if lines_k1 is None:
        sys.exit()
    elif lines_k1 == '':
        lines_k1 = '0'
    elif lines_k1 == 'all':
        lines_k1 = '1'
        for i in range(2, positions_quantity+1):
            lines_k1 = lines_k1 + ',' + str(i)

    lines_k1_data = lines_k1.replace('.', ',').split(',')
    lines_list_k1 = list(map(int, (lines_k1_data)))

    login()
    create_order()
    fill_order()
    fill_work_type()
    fill_lines_names()
    fill_emails()
    calculate()
    factorize()
    order_number = save_order_number()
    make_contract(order_number)
    save_contract(order_number)
