import openpyxl
import datetime

file_name = datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S' + '.xlsx')
wb = openpyxl.Workbook()
ws = wb.active

list1 = [1, 2, 10]

for i in range(1, 11):
    MATERIAL = '27001765'
    if i in list1:
        MATERIAL = '27001764'
    ws['A'+str(i)] = MATERIAL
    ws['B'+str(i)] = 1
    ws['E'+str(i)] = str(50150)

last_cell = ws.cell(ws.max_row, ws.max_column)
# print(last_cell.coordinate)
output = ws['A1':last_cell.coordinate]
# print(output)

wb.save(file_name)
